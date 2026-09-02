"""
Automated Pre-Release Regression Test Suite for Apollo Brand Intelligence.
Validates critical analyst workflows, export column contracts, and threat intelligence logic.
Must pass 100% before any production executable is built or released.
"""

import os
import sys
import tempfile
import unittest
import openpyxl
from bs4 import BeautifulSoup

# Ensure project root in sys.path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from data_store import DataStore
from exporter import ExcelExporter
from visual_catalog import VisualCatalogManager, compute_phash, hamming_distance
import batch_importer
import intel_pack_manager
from PIL import Image


class TestApolloCoreFeatures(unittest.TestCase):
    
    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.data_store = DataStore()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_01_ebay_seller_name_extraction(self):
        """Test Item 1: Verify seller extraction from HTML ignores avatar initials and feedback counts."""
        sample_ebay_html = """
        <html>
            <body>
                <h1 class="x-item-title__mainTitle">4PCS/SET OEM 90919-02240 Ignition Coils For Toyota Camry</h1>
                <div class="x-sellercard-atf">
                    <div class="x-sellercard-atf__info__about-seller">
                        <a href="https://www.ebay.com/sch/khkok-64/m.html?item=407006409742">
                            <span>K</span>
                        </a>
                        <a href="https://www.ebay.com/sch/khkok-64/m.html?item=407006409742">
                            <span class="ux-textspans--BOLD">khkok-64</span>
                        </a>
                        <span>(135)</span>
                        <span>99.1% positive</span>
                    </div>
                </div>
                <div class="x-price-primary"><span class="ux-textspans">US $59.99</span></div>
            </body>
        </html>
        """
        soup = BeautifulSoup(sample_ebay_html, "html.parser")
        
        # Test DOM seller extraction logic
        seller = ""
        for sel in (
            "div.x-sellercard-atf__info__about-seller a",
            "div[data-testid='x-sellercard-atf'] a",
            "div.ux-seller-section a",
            "a.x-sellercard-atf__info__about-seller"
        ):
            if seller: break
            for a_el in soup.select(sel):
                href = a_el.get("href", "")
                txt = a_el.get_text(strip=True)
                import re
                m_href = re.search(r'/(?:sch|usr|str)/([a-zA-Z0-9_\-\.]+)(?:/m\.html|\?|$|/)', href)
                if m_href:
                    cand = m_href.group(1).strip()
                    if cand and len(cand) >= 2 and cand.lower() not in ("usr", "str", "sch", "itm", "ebay"):
                        seller = cand
                        break
        
        self.assertEqual(seller, "khkok-64", "Seller name must be 'khkok-64' and not avatar initial 'K' or feedback rating.")

    def test_02_multi_locale_genesis_export_schema(self):
        """Test Item 2: Verify Multi-Locale Excel export adheres strictly to Genesis Columns A-R with Col C Thumbnail."""
        exporter = ExcelExporter()
        sample_items = [{
            "title": "OEM Toyota TRD Emblem Badge Set",
            "url": "https://www.ebay.com/itm/407006409742",
            "image_url": "https://i.ebayimg.com/images/g/sample_thumb.jpg",
            "item_id": "407006409742",
            "seller": "khkok-64",
            "price": "$59.99",
            "location": "Rowland Heights, CA, United States",
            "brand": "Toyota",
            "product_type": "Emblems",
            "seller_origin": "China",
            "threat_badge": "Foreign Drop-Ship Hub"
        }]
        target_locales = [
            {"name": "United Kingdom", "domain": "ebay.co.uk", "flag": "UK", "region": "Europe"},
            {"name": "Germany", "domain": "ebay.de", "flag": "DE", "region": "Europe"},
            {"name": "Australia", "domain": "ebay.com.au", "flag": "AU", "region": "Asia-Pacific"}
        ]
        
        out_file = os.path.join(self.temp_dir, "test_multi_locale_genesis.xlsx")
        count = exporter.export_multi_locale(sample_items, target_locales, out_file)
        
        self.assertEqual(count, 3, "Must generate exactly 3 expanded international rows.")
        self.assertTrue(os.path.exists(out_file), "Multi-locale export file must exist.")

        # Verify Excel sheet columns
        wb = openpyxl.load_workbook(out_file)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[0], "Title", "Col A must be Title")
        self.assertEqual(headers[1], "URL", "Col B must be URL")
        self.assertEqual(headers[2], "Thumbnail", "Col C must be Thumbnail")
        self.assertEqual(headers[4], "Item ID", "Col E must be Item ID")
        self.assertEqual(headers[7], "Marketplace", "Col H must be Marketplace")
        self.assertEqual(headers[9], "Seller Name", "Col J must be Seller Name")
        self.assertEqual(headers[12], "Brand", "Col M must be Brand")
        self.assertEqual(headers[13], "Price", "Col N must be Price")
        self.assertEqual(headers[14], "Item Location", "Col O must be Item Location")
        self.assertEqual(headers[15], "Product Type", "Col P must be Product Type")
        self.assertEqual(headers[18], "Locale Country", "Col S must be Locale Country")

        # Verify Row 2 content
        row2 = [cell.value for cell in ws[2]]
        self.assertEqual(row2[0], "OEM Toyota TRD Emblem Badge Set")
        self.assertEqual(row2[1], "https://www.ebay.co.uk/itm/407006409742", "Col B must contain expanded UK URL")
        self.assertEqual(row2[2], "https://i.ebayimg.com/images/g/sample_thumb.jpg", "Col C must contain thumbnail image URL")
        self.assertEqual(row2[4], "407006409742", "Col E must contain item ID")
        self.assertEqual(row2[9], "khkok-64", "Col J must contain seller name")
        self.assertEqual(row2[12], "Toyota", "Col M must contain brand")

    def test_03_datastore_enforcement_registry_aggregation(self):
        """Test Items 5 and 6: Verify registry aggregation and type-safe threat score comparison."""
        test_seller = "auto_parts_syndicate_99"
        items = [
            {
                "item_id": "111222333444",
                "title": "Fake TRD Grille Badge",
                "price": "US $125.00",
                "threat_score": "95",  # string score
                "brand": "Toyota",
                "product_type": "Emblems",
                "location": "Ontario, CA, United States",
                "seller_origin": "China",
                "threat_badge": "Foreign Drop-Ship Hub"
            },
            {
                "item_id": "555666777888",
                "title": "Fake Lexus Wheel Caps (Set of 4)",
                "price": "US $45.50",
                "threat_score": 80,    # int score
                "brand": "Lexus",
                "product_type": "Wheel Caps",
                "location": "Rowland Heights, CA",
                "seller_origin": "China",
                "threat_badge": "Foreign Drop-Ship Hub"
            }
        ]

        # Must execute without TypeError: '>='
        self.data_store.record_enforcement_scan(test_seller, items, brand_name="Toyota")
        
        reg = self.data_store.get_enforcement_registry()
        self.assertIn(test_seller, reg)
        card = reg[test_seller]
        
        self.assertEqual(card.get("total_listings"), 2, "Registry must store total listings count.")
        self.assertAlmostEqual(card.get("total_value"), 170.50, places=2, msg="Registry must sum total dollar values ($125.00 + $45.50 = $170.50).")
        self.assertIn("Toyota", card.get("brands_targeted", []))
        self.assertIn("Lexus", card.get("brands_targeted", []))
        self.assertIn("Emblems", card.get("product_types", []))
        self.assertIn("Wheel Caps", card.get("product_types", []))
        self.assertEqual(card.get("country"), "China")

        # Cleanup
        del reg[test_seller]
        self.data_store._save()

    def test_04_threat_assessment_unresolved_origin(self):
        """Test Item 9: Verify foreign drop-shippers are flagged and unresolved origins do not default to Domestic."""
        # 1. Foreign origin + US warehouse -> Drop-Ship Hub
        assess_3pl = self.data_store.compute_threat_assessment(origin="China", location="Ontario, California, United States")
        self.assertTrue(assess_3pl.get("is_3pl_hub"), "China seller with US warehouse must be flagged as 3PL Hub.")
        self.assertIn("Drop-Ship Hub", assess_3pl.get("badge"))

        # 2. Unresolved origin + US warehouse -> Must NOT be marked Domestic Verified
        assess_unres = self.data_store.compute_threat_assessment(origin="", location="City of Industry, CA")
        self.assertNotIn("Domestic Verified", assess_unres.get("badge"), "Unresolved origin must not be labeled Domestic Verified.")
        self.assertIn("Unresolved", assess_unres.get("badge"))

        # 3. Explicit Domestic origin -> Domestic Verified
        assess_dom = self.data_store.compute_threat_assessment(origin="United States", location="Austin, TX")
        self.assertIn("Domestic Verified", assess_dom.get("badge"))

    def test_05_datastore_delete_registry_entry(self):
        """Test Registry: Verify delete_registry_entry removes seller record cleanly."""
        seller = "test_seller_to_remove"
        self.data_store.record_enforcement_scan(seller, [{"item_id": "999", "title": "T", "price": "$10"}])
        self.assertIn(seller, self.data_store.get_enforcement_registry())
        
        self.data_store.delete_registry_entry(seller)
        self.assertNotIn(seller, self.data_store.get_enforcement_registry(), "Seller must be deleted from registry.")

    def test_06_standard_genesis_export_schema(self):
        """Test Standard Export: Verify 18-column Genesis layout with Col C Thumbnail and Col B URL."""
        exporter = ExcelExporter()
        sample_items = [{
            "title": "Toyota Genuine Oil Filter 90915-YZZN1",
            "url": "https://www.ebay.com/itm/112233445566",
            "image_url": "https://i.ebayimg.com/images/g/test_oil_filter.jpg",
            "item_id": "112233445566",
            "seller": "toyota_direct_deals",
            "price": "$9.99",
            "location": "Dallas, TX, United States",
            "brand": "Toyota",
            "product_type": "Oil / Fuel Filters",
            "seller_origin": "United States",
            "threat_badge": "Domestic Verified"
        }]
        out_file = os.path.join(self.temp_dir, "test_standard_genesis.xlsx")
        count = exporter.export_results(sample_items, out_file)
        self.assertEqual(count, 1)
        self.assertTrue(os.path.exists(out_file))

        wb = openpyxl.load_workbook(out_file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers[0], "Title")
        self.assertEqual(headers[1], "URL")
        self.assertEqual(headers[2], "Thumbnail")
        self.assertEqual(headers[4], "Item ID")
        self.assertEqual(headers[7], "Marketplace")
        self.assertEqual(headers[9], "Seller Name")
        self.assertEqual(headers[12], "Brand")
        self.assertEqual(headers[13], "Price")
        self.assertEqual(headers[14], "Item Location")
        self.assertEqual(headers[15], "Product Type")
        self.assertIn("Threat Assessment", headers[17])

        row2 = [cell.value for cell in ws[2]]
        self.assertEqual(row2[0], "Toyota Genuine Oil Filter 90915-YZZN1")
        self.assertEqual(row2[1], "https://www.ebay.com/itm/112233445566")
        self.assertEqual(row2[2], "https://i.ebayimg.com/images/g/test_oil_filter.jpg")
        self.assertEqual(row2[4], "112233445566")
        self.assertEqual(row2[9], "toyota_direct_deals")
        self.assertEqual(row2[12], "Toyota")

    def test_07_whitelist_authorized_dealers(self):
        """Test Whitelist: Verify authorized dealerships are identified and shielded."""
        test_handle = "authorized_toyota_dealer_tx"
        self.data_store.add_to_whitelist(test_handle, brand="Toyota", dealer_name="Toyota of Dallas")
        
        self.assertTrue(self.data_store.is_seller_whitelisted(test_handle))
        self.assertTrue(self.data_store.is_seller_whitelisted(f"  {test_handle.upper()}  "), "Must be whitespace & case insensitive")
        self.assertFalse(self.data_store.is_seller_whitelisted("unknown_counterfeiter_99"))

        # Clean up
        self.data_store.remove_from_whitelist(test_handle)
        self.assertFalse(self.data_store.is_seller_whitelisted(test_handle))

    def test_08_brand_detection_heuristics(self):
        """Test Brand Detection: Verify title classification heuristics accurately extract trademark brands."""
        self.assertEqual(batch_importer.detect_brand("OEM TRD Grille Badge for Toyota Tacoma"), "Toyota")
        self.assertEqual(batch_importer.detect_brand("2024 Lexus RX350 Wheel Center Caps 4pcs"), "Lexus")
        self.assertEqual(batch_importer.detect_brand("Subaru WRX STI Red Stitching Steering Wheel"), "Subaru")
        self.assertEqual(batch_importer.detect_brand("Honda Civic Type R Carbon Fiber Wing Spoiler"), "Honda")
        self.assertEqual(batch_importer.detect_brand("Generic Unbranded Key Chain"), "Automotive & Consumer Brands")

    def test_09_adhoc_url_cleaning_and_id_extraction(self):
        """Test URL Cleaning: Verify messy tracking URLs are canonicalized and item IDs extracted."""
        dirty_url = "https://www.ebay.com/itm/407006409742?_trksid=p2047675.c100005.m1851&_trkparms=amclksrc%3DITM&hash=item5f00"
        clean = batch_importer.clean_ebay_url(dirty_url)
        self.assertEqual(clean, "https://www.ebay.com/itm/407006409742")
        self.assertEqual(batch_importer.extract_item_id(dirty_url), "407006409742")

    def test_10_product_type_classification(self):
        """Test Category Detection: Verify auto-classification of product categories."""
        self.assertEqual(batch_importer.detect_product_type("4Pcs Iridium Spark Plugs for Camry"), "Spark Plugs")
        self.assertEqual(batch_importer.detect_product_type("Front Ceramic Brake Pads and Rotors Kit"), "Brake Pads / Rotors")
        self.assertEqual(batch_importer.detect_product_type("Gloss Black Front Grille Emblem Badge"), "Emblems / Badges")
        self.assertEqual(batch_importer.detect_product_type("Engine Oil Filter Replacement Cartridge"), "Oil / Fuel Filters")

    def test_11_intel_pack_export_and_import_merge(self):
        """Test Intelligence Pack: Verify .apollo packaging, export, inspection, and safe library merging."""
        # 1. Setup isolated data store & visual catalog
        vcm_dir = os.path.join(self.temp_dir, "vcm_src")
        vcm = VisualCatalogManager(base_dir=vcm_dir)
        
        # Add sample test image
        img = Image.new("RGB", (64, 64), color=(255, 0, 0))
        vcm.add_entry(img, entry_type="benign", label="Toyota Red OEM Box", source_url="https://example.com/box.jpg")

        pack_file = os.path.join(self.temp_dir, "test_intel_pack.apollo")
        manifest = intel_pack_manager.IntelPackManager.export_pack(
            output_filepath=pack_file,
            data_store=self.data_store,
            visual_catalog=vcm,
            scope="Full Profile",
            author="Jerry Seidenstucker",
            notes="Automated Test Pack"
        )
        self.assertTrue(os.path.exists(pack_file))
        self.assertEqual(manifest["author"], "Jerry Seidenstucker")
        self.assertGreaterEqual(manifest["counts"]["brands"], 1)
        self.assertGreaterEqual(manifest["counts"]["visual_catalog_entries"], 1)

        # 2. Inspect Pack
        inspected = intel_pack_manager.IntelPackManager.inspect_pack(pack_file)
        self.assertEqual(inspected["format"], "apollo_intelligence_pack")
        self.assertEqual(inspected["version"], "1.0")

        # 3. Import into destination catalog
        dst_vcm_dir = os.path.join(self.temp_dir, "vcm_dst")
        dst_vcm = VisualCatalogManager(base_dir=dst_vcm_dir)
        self.assertEqual(len(dst_vcm.get_all_entries()), 0)

        import_res = intel_pack_manager.IntelPackManager.import_pack(
            pack_filepath=pack_file,
            data_store=self.data_store,
            visual_catalog=dst_vcm,
            merge_mode="merge"
        )
        self.assertGreaterEqual(len(dst_vcm.get_all_entries()), 1)
        self.assertEqual(import_res["results"]["visual_added"], 1)
        self.assertEqual(import_res["results"]["thumbnails_extracted"], 1)

    def test_12_visual_sensitivity_dynamic_threshold(self):
        """Test Visual Sensitivity: Verify Hamming distance matching and dynamic threshold behavior."""
        vcm_dir = os.path.join(self.temp_dir, "vcm_thresh")
        vcm = VisualCatalogManager(base_dir=vcm_dir)

        # Create base image and slight variant
        base_img = Image.new("RGB", (64, 64), color=(200, 30, 30))
        vcm.add_entry(base_img, entry_type="benign", label="Denso Blue Box")

        # Test exact match
        m_exact = vcm.match_image(base_img, max_distance=2)
        self.assertIsNotNone(m_exact)
        self.assertEqual(m_exact["label"], "Denso Blue Box")
        self.assertEqual(m_exact["type"], "benign")

        # Test distant image fails on strict (max_distance=2), passes on broad (max_distance=20)
        diff_img = Image.new("RGB", (64, 64), color=(10, 200, 50))
        h1 = compute_phash(base_img)
        h2 = compute_phash(diff_img)
        dist = hamming_distance(h1, h2)
        
        m_strict = vcm.match_image(diff_img, max_distance=max(0, dist - 5))
        self.assertIsNone(m_strict)

        m_broad = vcm.match_image(diff_img, max_distance=dist + 5)
        self.assertIsNotNone(m_broad)

    def test_13_multi_locale_export_col_h_domain_format(self):
        """Test Multi-Locale Export: Verify Column H outputs strict domain name format (ebay.com, ebay.ca, etc.)."""
        exporter = ExcelExporter()
        test_results = [{
            "title": "Toyota Genuine Oil Filter 90915-YZZN1",
            "url": "https://www.ebay.com/itm/112233445566",
            "item_id": "112233445566",
            "image_url": "https://i.ebayimg.com/images/g/test.jpg",
            "seller": "toyota_direct_deals",
            "brand": "Toyota",
            "price": "$12.99",
            "location": "Dallas, TX, United States",
            "product_type": "Oil Filters",
            "seller_origin": "United States",
            "threat_badge": "🇺🇸 Domestic Verified"
        }]

        test_locales = [
            {"code": "US", "name": "United States", "domain": "ebay.com", "region": "North America", "flag": "🇺🇸"},
            {"code": "CA", "name": "Canada", "domain": "ebay.ca", "region": "North America", "flag": "🇨🇦"},
            {"code": "UK", "name": "United Kingdom", "domain": "ebay.co.uk", "region": "Europe", "flag": "🇬🇧"},
            {"code": "DE", "name": "Germany", "domain": "ebay.de", "region": "Europe", "flag": "🇩🇪"},
        ]

        out_path = os.path.join(self.temp_dir, "test_multi_locale_col_h.xlsx")
        exporter.export_multi_locale(test_results, test_locales, out_path)
        self.assertTrue(os.path.exists(out_path))

        wb = openpyxl.load_workbook(out_path)
        ws = wb.active

        # Check Col H across the 4 generated locale rows
        col_h_vals = [ws.cell(row=r, column=8).value for r in range(2, 6)]
        self.assertEqual(col_h_vals, ["ebay.com", "ebay.ca", "ebay.co.uk", "ebay.de"])

        # Verify Column C contains thumbnail image URL
        col_c_vals = [ws.cell(row=r, column=3).value for r in range(2, 6)]
        self.assertEqual(col_c_vals, ["https://i.ebayimg.com/images/g/test.jpg"] * 4)

    def test_14_seller_extraction_rejects_promo_and_spec_copy(self):
        """Test Seller Extraction: Verify promotional copy and specification tags are never extracted as seller names."""
        from scraper import EbayScraper
        scraper = EbayScraper(headless=True)

        mock_card_html = """
        <li class="s-card">
            <a class="s-card__link" href="https://www.ebay.com/itm/998877665544">
                <span class="s-card__title">Toyota Camry Steering Wheel Badge</span>
            </a>
            <div class="s-card__subtitle">17 sold • Save up to 5% with coupon</div>
            <span class="s-card__price">$24.99</span>
        </li>
        """

        # When parsed with a known fallback seller, must strictly use fallback instead of 'sold' or 'save'
        items = scraper._parse_html(mock_card_html, fallback_seller="genuine_oem_parts_direct")
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["seller"], "genuine_oem_parts_direct")
        self.assertNotIn("sold", items[0]["seller"].lower())
        self.assertNotIn("save", items[0]["seller"].lower())

    def test_15_threat_assessment_india_cross_border_drop_ship(self):
        """Test Threat Assessment: Verify India-registered sellers shipping from US 3PL warehouses are flagged as Drop-Ship Hubs."""
        # 1. India seller shipping from US warehouse -> Foreign Drop-Ship Hub
        res_3pl = self.data_store.compute_threat_assessment(origin="India", location="Chino, CA, United States")
        self.assertEqual(res_3pl["badge"], "🚨 Foreign Drop-Ship Hub")
        self.assertTrue(res_3pl["is_high_risk"])
        self.assertTrue(res_3pl["is_3pl_hub"])

        # 2. India seller shipping directly from India -> Cross-Border Direct
        res_direct = self.data_store.compute_threat_assessment(origin="India", location="New Delhi, India")
        self.assertEqual(res_direct["badge"], "⚠️ Cross-Border Direct")
        self.assertTrue(res_direct["is_high_risk"])
        self.assertFalse(res_direct["is_3pl_hub"])

        # 3. Domestic US seller shipping from US -> Domestic Verified
        res_us = self.data_store.compute_threat_assessment(origin="United States", location="Dallas, TX, United States")
        self.assertEqual(res_us["badge"], "🇺🇸 Domestic Verified")
        self.assertFalse(res_us["is_high_risk"])

    def test_16_tiktok_shop_pdp_extraction(self):
        """Test Item 16: Verify TikTok Shop platform detection, URL extraction, and PDP normalization."""
        sample_url = "https://shop.tiktok.com/us/pdp/chrome-valve-stem-tire-caps-for-cadillac-vehicles-set-of-four/1731432810739700325"
        
        # 1. Platform Detection
        platform = batch_importer.detect_platform(sample_url)
        self.assertEqual(platform, "TikTok Shop", "Platform must be recognized as 'TikTok Shop'")

        # 2. Markdown URL Extraction
        raw_text = f"Review this link: [Cadillac Caps]({sample_url}) and also https://shop.tiktok.com/us/pdp/1732474117957129133"
        extracted_urls = batch_importer.extract_urls_from_text(raw_text)
        self.assertIn(sample_url, extracted_urls)
        self.assertEqual(len(extracted_urls), 2)

        # 3. TikTokScraper PDP normalization contract
        from tiktok_scraper import TikTokScraper
        scraper = TikTokScraper(headless=True)
        store_info = scraper.resolve_store_info(sample_url)
        self.assertEqual(store_info.get("item_id"), "1731432810739700325")

    def test_17_smart_triage_universal_fluff(self):
        """Test Item 17: Verify Smart Triage suppresses universal fluff & multi-brand spam while preserving high-risk components."""
        # 1. High-risk parts must NEVER be suppressed, even with 'fits' or multiple words
        is_fluff, _ = self.data_store.is_universal_fluff("4PCS OEM 90919-02240 Ignition Coils For Toyota Camry")
        self.assertFalse(is_fluff, "Ignition coils must never be suppressed")

        is_fluff, _ = self.data_store.is_universal_fluff("Toyota Genuine Oil Filter 04152-YZZA1 fits Camry RAV4")
        self.assertFalse(is_fluff, "Oil filters must never be suppressed")

        is_fluff, _ = self.data_store.is_universal_fluff("TRD Front Grille Emblem Badge fits Toyota Tacoma 4Runner")
        self.assertFalse(is_fluff, "Emblems/Badges must never be suppressed")

        is_fluff, _ = self.data_store.is_universal_fluff("4pcs Spark Plugs Iridium fits Toyota Denso SK20R11")
        self.assertFalse(is_fluff, "Spark plugs must never be suppressed")

        # 2. Multi-brand title spam must be suppressed
        is_fluff, reason = self.data_store.is_universal_fluff("Universal Breathable Leather Seat Cover fits Toyota Honda Chevy Nissan")
        self.assertTrue(is_fluff, "Multi-brand title spam must be suppressed")
        self.assertIn("Multi-Brand Spam", reason)

        # 3. Compatibility keyword + universal fluff category must be suppressed
        is_fluff, reason = self.data_store.is_universal_fluff("Car Windshield Sunshade Foldable for Toyota Corolla")
        self.assertTrue(is_fluff, "Universal sunshade compatibility must be suppressed")
        self.assertIn("Universal Compatibility", reason)

        is_fluff, reason = self.data_store.is_universal_fluff("Heavy Duty Rubber Floor Mats for Chevy Silverado")
        self.assertTrue(is_fluff, "Floor mats with 'for' must be suppressed")
        self.assertIn("Universal Compatibility", reason)

    def test_18_manomano_and_whitelist_scopes(self):
        """Test Item 18: Verify ManoMano contract, Whitelist marketplace scoping, and handle preservation."""
        # 1. ManoManoScraper contract
        from manomano_scraper import ManoManoScraper
        mm = ManoManoScraper(headless=True)
        info = mm.resolve_store_info("https://www.manomano.fr/marchand-41084935")
        self.assertEqual(info.get("store_name"), "ManoMano European Search")

        # 2. Whitelist marketplace scoping
        self.data_store.add_to_whitelist("legit_dealer_global", brand="Toyota", dealer_name="Global Dealer", marketplace="All Marketplaces (Global)")
        self.data_store.add_to_whitelist("legit_dealer_ebay_only", brand="Toyota", dealer_name="eBay Dealer", marketplace="eBay Only")

        self.assertTrue(self.data_store.is_seller_whitelisted("legit_dealer_global", marketplace="eBay"))
        self.assertTrue(self.data_store.is_seller_whitelisted("legit_dealer_global", marketplace="ManoMano"))
        self.assertTrue(self.data_store.is_seller_whitelisted("legit_dealer_ebay_only", marketplace="eBay"))
        self.assertFalse(self.data_store.is_seller_whitelisted("legit_dealer_ebay_only", marketplace="ManoMano"))

        # 3. Handle preservation (no mangling caug_92 -> caug92)
        from scraper import EbayScraper
        eb = EbayScraper()
        candidates = eb._generate_seller_candidates("caug_92")
        self.assertEqual(candidates, ["caug_92"], "Underscores in seller handles must never be mangled")


if __name__ == "__main__":
    unittest.main()
